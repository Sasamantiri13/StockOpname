import openpyxl
import sys

def verify_excel_calculation():
    """Verify that Excel Total Variance Value matches web application"""
    
    try:
        # Load the Excel file
        wb = openpyxl.load_workbook("Stock_Opname_DSS_Template.xlsx", data_only=True)
        
        print("=== Verification of Excel Total Variance Value ===\n")
        
        # Check DSS-SPK_Analysis sheet for individual variance values
        analysis_sheet = wb['DSS-SPK_Analysis']
        
        print("Individual Variance Values from Excel:")
        variance_values = []
        for row in range(2, 12):  # Check rows 2-11 for data
            product_code = analysis_sheet.cell(row=row, column=1).value
            if product_code:  # If there's a product code, get the variance value
                variance_value = analysis_sheet.cell(row=row, column=8).value  # Column H (Variance_Value)
                if variance_value is not None:
                    print(f"Row {row} ({product_code}): {variance_value:,}")
                    variance_values.append(abs(variance_value))
        
        manual_total = sum(variance_values)
        print(f"\nManual calculation (sum of absolute values): {manual_total:,}")
        
        # Check Summary Dashboard for the Total Variance Value formula result
        summary_sheet = wb['Summary_Dashboard']
        
        print("\nChecking Summary Dashboard...")
        total_variance_from_excel = None
        for row in range(1, 20):
            for col in range(1, 10):
                cell_value = summary_sheet.cell(row=row, column=col).value
                if cell_value and 'Total Variance' in str(cell_value):
                    formula_cell = summary_sheet.cell(row=row, column=col+1)
                    total_variance_from_excel = formula_cell.value
                    print(f"Found Total Variance Value: {total_variance_from_excel}")
                    break
            if total_variance_from_excel is not None:
                break
        
        # Expected result from web application
        expected_total = 17750000  # Rp 17,750,000
        
        print(f"\n=== COMPARISON ===")
        print(f"Expected (Web App): Rp {expected_total:,}")
        print(f"Manual calculation: Rp {manual_total:,}")
        if total_variance_from_excel is not None:
            print(f"Excel formula result: Rp {total_variance_from_excel:,}")
            
            if abs(total_variance_from_excel - expected_total) < 1:
                print("✅ SUCCESS: Excel calculation matches web application!")
                return True
            else:
                print("❌ MISMATCH: Excel calculation does not match web application")
                print(f"Difference: Rp {abs(total_variance_from_excel - expected_total):,}")
                return False
        else:
            print("❌ ERROR: Could not find Total Variance Value in Excel Summary Dashboard")
            return False
            
    except Exception as e:
        print(f"Error: {e}")
        return False

if __name__ == "__main__":
    success = verify_excel_calculation()
    sys.exit(0 if success else 1)

