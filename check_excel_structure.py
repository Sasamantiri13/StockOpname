import pandas as pd
import openpyxl

# Load the Excel file
excel_file = "Stock_Opname_DSS_Template.xlsx"
wb = openpyxl.load_workbook(excel_file, data_only=False)

print("=== Excel File Structure ===")
print("Sheet names:", wb.sheetnames)
print()

# Check DSS_Analysis sheet (might be named differently)
analysis_sheet = None
for sheet_name in wb.sheetnames:
    if 'Analysis' in sheet_name or 'DSS' in sheet_name:
        analysis_sheet = wb[sheet_name]
        print(f"Found analysis sheet: {sheet_name}")
        break

if analysis_sheet:
    print(f"\n=== {analysis_sheet.title} Sheet Structure ===")
    
    # Check headers
    headers = []
    for col in range(1, 20):  # Check first 20 columns
        cell_value = analysis_sheet.cell(row=1, column=col).value
        if cell_value:
            headers.append(cell_value)
        else:
            break
    
    print("Headers:", headers)
    
    # Check Total Variance Value formula
    print("\n=== Total Variance Value Formula ===")
    
    # Look for Total Variance Value in Summary sheet
    summary_sheet = wb['Summary_Dashboard']
    for row in range(1, 20):
        for col in range(1, 10):
            cell = summary_sheet.cell(row=row, column=col)
            if cell.value and 'Total Variance' in str(cell.value):
                print(f"Found 'Total Variance' at row {row}, col {col}: {cell.value}")
                # Check the formula in the next column
                formula_cell = summary_sheet.cell(row=row, column=col+1)
                print(f"Formula: {formula_cell.value}")
                if hasattr(formula_cell, 'formula') and formula_cell.formula:
                    print(f"Excel Formula: {formula_cell.formula}")
                break
    
    # Check sample data and formulas in analysis sheet
    print(f"\n=== Sample Data in {analysis_sheet.title} ===")
    for row in range(2, 4):  # Check first 2 data rows
        variance_value_col = None
        for col in range(1, len(headers) + 1):
            header = analysis_sheet.cell(row=1, column=col).value
            if header and 'Variance_Value' in str(header):
                variance_value_col = col
                break
        
        if variance_value_col:
            cell = analysis_sheet.cell(row=row, column=variance_value_col)
            print(f"Row {row} Variance Value: {cell.value}")
            if hasattr(cell, 'formula') and cell.formula:
                print(f"Row {row} Formula: {cell.formula}")

else:
    print("No analysis sheet found!")

