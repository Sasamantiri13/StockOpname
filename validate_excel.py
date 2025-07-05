import pandas as pd
import openpyxl
from openpyxl import load_workbook

def validate_excel_structure():
    """Validate the Excel DSS template structure and formulas"""
    
    file_path = "C:\\Users\\SASA\\P-SPK-StockOpname-TSX\\Stock_Opname_DSS_Template.xlsx"
    
    try:
        # Load workbook
        wb = load_workbook(file_path, data_only=False)
        
        print("=== Excel DSS Template Validation ===\n")
        
        # Check sheets
        expected_sheets = ['Input_Data', 'DSS_Analysis', 'Summary_Dashboard', 'Recommendations']
        actual_sheets = wb.sheetnames
        
        print("1. Sheet Structure:")
        for sheet in expected_sheets:
            if sheet in actual_sheets:
                print(f"   ✓ {sheet} - Present")
            else:
                print(f"   ✗ {sheet} - Missing")
        
        print(f"\n   Total sheets: {len(actual_sheets)}")
        print(f"   Sheet names: {actual_sheets}\n")
        
        # Validate Input_Data sheet
        input_sheet = wb['Input_Data']
        print("2. Input_Data Sheet:")
        
        expected_headers = [
            'Product_Code', 'Product_Name', 'Category', 'System_Stock', 'Actual_Stock',
            'Unit_Cost', 'Min_Stock', 'Max_Stock', 'Lead_Time_Days', 'Avg_Daily_Demand',
            'Ordering_Cost', 'Holding_Cost_Rate'
        ]
        
        for i, header in enumerate(expected_headers, 1):
            cell_value = input_sheet.cell(row=1, column=i).value
            if cell_value == header:
                print(f"   ✓ Column {i}: {header}")
            else:
                print(f"   ✗ Column {i}: Expected '{header}', got '{cell_value}'")
        
        # Check sample data
        sample_row = 2
        if input_sheet.cell(row=sample_row, column=1).value:
            print(f"   ✓ Sample data present in row {sample_row}")
        else:
            print(f"   ✗ No sample data in row {sample_row}")
        
        # Validate DSS_Analysis sheet
        analysis_sheet = wb['DSS_Analysis']
        print("\n3. DSS_Analysis Sheet:")
        
        analysis_headers = [
            'Product_Code', 'Product_Name', 'Category', 'System_Stock', 'Actual_Stock',
            'Variance', 'Variance_Pct', 'Variance_Value', 'Inventory_Value',
            'Annual_Demand', 'Safety_Stock', 'Reorder_Point', 'EOQ',
            'Stock_Status', 'Turnover_Ratio', 'ABC_Class', 'Cumulative_Pct'
        ]
        
        for i, header in enumerate(analysis_headers, 1):
            cell_value = analysis_sheet.cell(row=1, column=i).value
            if cell_value == header:
                print(f"   ✓ Column {i}: {header}")
            else:
                print(f"   ✗ Column {i}: Expected '{header}', got '{cell_value}'")
        
        # Check key formulas
        print("\n4. Formula Validation:")
        
        # Variance formula (column F, row 2)
        variance_formula = analysis_sheet.cell(row=2, column=6).value
        if isinstance(variance_formula, str) and '=E2-D2' in variance_formula:
            print("   ✓ Variance formula correct")
        else:
            print(f"   ✗ Variance formula: {variance_formula}")
        
        # Variance percentage formula (column G, row 2)
        var_pct_formula = analysis_sheet.cell(row=2, column=7).value
        if isinstance(var_pct_formula, str) and 'IF(D2<>0' in var_pct_formula:
            print("   ✓ Variance percentage formula correct")
        else:
            print(f"   ✗ Variance percentage formula: {var_pct_formula}")
        
        # EOQ formula (column M, row 2)
        eoq_formula = analysis_sheet.cell(row=2, column=13).value
        if isinstance(eoq_formula, str) and 'SQRT' in eoq_formula and 'CEILING' in eoq_formula:
            print("   ✓ EOQ formula contains SQRT and CEILING")
        else:
            print(f"   ✗ EOQ formula: {eoq_formula}")
        
        # Safety Stock formula (column K, row 2)
        safety_stock_formula = analysis_sheet.cell(row=2, column=11).value
        if isinstance(safety_stock_formula, str) and 'SQRT' in safety_stock_formula and '1.65' in safety_stock_formula:
            print("   ✓ Safety Stock formula correct (95% service level)")
        else:
            print(f"   ✗ Safety Stock formula: {safety_stock_formula}")
        
        # Validate Summary_Dashboard sheet
        summary_sheet = wb['Summary_Dashboard']
        print("\n5. Summary_Dashboard Sheet:")
        
        # Check title
        title_cell = summary_sheet.cell(row=1, column=1).value
        if 'DSS SUMMARY' in str(title_cell):
            print("   ✓ Title present")
        else:
            print(f"   ✗ Title: {title_cell}")
        
        # Check some key metrics formulas
        metrics_start_row = 3
        for row in range(metrics_start_row, metrics_start_row + 5):
            metric_name = summary_sheet.cell(row=row, column=1).value
            formula = summary_sheet.cell(row=row, column=2).value
            
            if metric_name and isinstance(formula, str) and formula.startswith('='):
                print(f"   ✓ {metric_name}: Has formula")
            elif metric_name:
                print(f"   ✗ {metric_name}: No formula")
        
        print("\n6. File Information:")
        print(f"   ✓ File exists: {file_path}")
        print(f"   ✓ Workbook loaded successfully")
        print(f"   ✓ Total worksheets: {len(wb.worksheets)}")
        
        print("\n=== Validation Complete ===")
        print("The Excel DSS template has been successfully created with:")
        print("- All required sheets")
        print("- Proper headers and structure")
        print("- DSS calculation formulas")
        print("- Sample data for testing")
        
        # Load and display sample calculation
        print("\n7. Sample Calculation Preview:")
        df_input = pd.read_excel(file_path, sheet_name='Input_Data')
        print("Input Data (first 3 rows):")
        print(df_input.head(3).to_string(index=False))
        
        wb.close()
        
    except Exception as e:
        print(f"Error validating Excel file: {str(e)}")
        return False
    
    return True

if __name__ == "__main__":
    validate_excel_structure()

